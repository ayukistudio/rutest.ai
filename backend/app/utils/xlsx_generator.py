# app/utils/xlsx_generator.py

from io import BytesIO
from typing import List, Dict, Any
from datetime import datetime
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter


def generate_xlsx_bytes(test_cases: List[Dict[str, Any]]) -> bytes:
    wb: Workbook = Workbook()
    ws = wb.active
    ws.title = "Тест-кейсы"

    # === СТИЛИ ===
    # Шрифты
    header_font = Font(bold=True, size=11, color="FFFFFF", name="Calibri")
    field_label_font = Font(bold=True, size=10, color="1F4E78", name="Calibri")
    normal_font = Font(size=10, name="Calibri")
    number_font = Font(bold=True, size=11, color="1F4E78", name="Calibri")
    
    # Выравнивание
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_top_align = Alignment(horizontal="left", vertical="top", wrap_text=True)
    left_center_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    
    # Границы
    thin_side = Side(border_style="thin", color="B4C7E7")
    medium_side = Side(border_style="medium", color="4472C4")
    border_thin = Border(top=thin_side, left=thin_side, right=thin_side, bottom=thin_side)
    border_thick_bottom = Border(top=thin_side, left=thin_side, right=thin_side, bottom=medium_side)
    
    # Заливки
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    number_fill = PatternFill(start_color="E7E6F7", end_color="E7E6F7", fill_type="solid")
    title_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    precondition_fill = PatternFill(start_color="E2F0D9", end_color="E2F0D9", fill_type="solid")
    steps_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    expected_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    comment_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    field_label_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    field_value_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    # === ЗАГОЛОВОК ===
    headers = ["№", "Название", "Предусловие", "Шаги", "Ожидаемый результат", "Комментарий"]
    ws.append(headers)
    ws.row_dimensions[1].height = 40

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.alignment = header_align
        cell.fill = header_fill
        cell.border = border_thin

    # === ШИРИНА КОЛОНОК ===
    column_widths = {
        'A': 5,   # Номер
        'B': 30,  # Название
        'C': 28,  # Предусловие
        'D': 40,  # Шаги
        'E': 40,  # Ожидаемый результат
        'F': 25   # Комментарий
    }
    
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width

    # Закрепление заголовка
    ws.freeze_panes = "A2"

    current_row = 2

    # === ГЕНЕРАЦИЯ ТЕСТ-КЕЙСОВ ===
    for idx, raw_tc in enumerate(test_cases, start=1):
        # Нормализация данных
        tc = {
            "title": raw_tc.get("title", "") or raw_tc.get("name", "") or "",
            "preconditions": raw_tc.get("preconditions", "") or raw_tc.get("setup", "") or "",
            "steps": raw_tc.get("steps", []),
            "expectedResult": raw_tc.get("expectedResult", "") or raw_tc.get("expected", "") or "",
            "description": raw_tc.get("description", "") or raw_tc.get("desc", "") or "",
            "context": raw_tc.get("context", "") or raw_tc.get("extra", "") or "",
            "result": raw_tc.get("result", "") or "",
            "priority": raw_tc.get("priority") or "Не указано",
        }

        # Форматирование шагов
        steps_text = _format_steps(tc["steps"])
        expected_text = _format_expected(tc["steps"])

        # === ОСНОВНАЯ СТРОКА ===
        ws.append([
            idx,
            tc["title"],
            tc["preconditions"],
            steps_text,
            expected_text,
            tc["expectedResult"]
        ])
        
        main_row = current_row
        
        # Стили для основной строки
        # Номер
        cell_num = ws.cell(row=main_row, column=1)
        cell_num.font = number_font
        cell_num.alignment = center_align
        cell_num.fill = number_fill
        cell_num.border = border_thin
        
        # Название
        cell_title = ws.cell(row=main_row, column=2)
        cell_title.font = Font(bold=True, size=10, name="Calibri")
        cell_title.alignment = left_top_align
        cell_title.fill = title_fill
        cell_title.border = border_thin
        
        # Предусловие
        cell_pre = ws.cell(row=main_row, column=3)
        cell_pre.font = normal_font
        cell_pre.alignment = left_top_align
        cell_pre.fill = precondition_fill
        cell_pre.border = border_thin
        
        # Шаги
        cell_steps = ws.cell(row=main_row, column=4)
        cell_steps.font = normal_font
        cell_steps.alignment = left_top_align
        cell_steps.fill = steps_fill
        cell_steps.border = border_thin
        
        # Ожидаемый результат
        cell_exp = ws.cell(row=main_row, column=5)
        cell_exp.font = normal_font
        cell_exp.alignment = left_top_align
        cell_exp.fill = expected_fill
        cell_exp.border = border_thin
        
        # Комментарий
        cell_com = ws.cell(row=main_row, column=6)
        cell_com.font = normal_font
        cell_com.alignment = left_top_align
        cell_com.fill = comment_fill
        cell_com.border = border_thin
        
        # Высота строки
        ws.row_dimensions[main_row].height = _calculate_row_height([
            tc["title"],
            tc["preconditions"],
            steps_text,
            expected_text,
            tc["expectedResult"]
        ])
        
        current_row += 1

        # === ДОПОЛНИТЕЛЬНЫЕ ПОЛЯ ===
        additional_fields = [
            ("Описание:", tc["description"]),
            ("Контекст:", tc["context"]),
            ("Приоритет:", tc["priority"]),
            ("Результат:", tc["result"])
        ]

        for field_label, field_value in additional_fields:
            if field_value and str(field_value).strip():
                ws.append(["", field_label, field_value, "", "", ""])
                field_row = current_row
                
                # Объединение ячеек для значения
                ws.merge_cells(start_row=field_row, start_column=3, end_row=field_row, end_column=6)
                
                # Пустая ячейка номера
                ws.cell(row=field_row, column=1).border = border_thin
                ws.cell(row=field_row, column=1).fill = number_fill
                
                # Метка поля
                label_cell = ws.cell(row=field_row, column=2)
                label_cell.font = field_label_font
                label_cell.alignment = left_center_align
                label_cell.fill = field_label_fill
                label_cell.border = border_thin
                
                # Значение поля
                value_cell = ws.cell(row=field_row, column=3)
                value_cell.font = normal_font
                value_cell.alignment = left_top_align
                value_cell.fill = field_value_fill
                value_cell.border = border_thin
                
                # Высота строки
                ws.row_dimensions[field_row].height = min(max(15 * len(str(field_value).splitlines()), 20), 120)
                
                current_row += 1

        # === РАЗДЕЛИТЕЛЬ ===
        ws.append(["", "", "", "", "", ""])
        sep_row = current_row
        
        for col_idx in range(1, 7):
            sep_cell = ws.cell(row=sep_row, column=col_idx)
            sep_cell.border = border_thick_bottom
        
        ws.row_dimensions[sep_row].height = 8
        current_row += 1

    # === СОХРАНЕНИЕ ===
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio.read()


def _format_steps(steps_list: List[Dict[str, Any]]) -> str:
    """Форматирует список шагов с нумерацией."""
    if not steps_list:
        return ""
    
    lines = []
    for i, step in enumerate(steps_list):
        step_no = step.get("step") if step.get("step") is not None else i + 1
        action = step.get("action") or step.get("description") or ""
        if action:
            lines.append(f"{step_no}. {action}")
    
    return "\n\n".join(lines)


def _format_expected(steps_list: List[Dict[str, Any]]) -> str:
    """Форматирует ожидаемые результаты для шагов."""
    if not steps_list:
        return ""
    
    lines = []
    for i, step in enumerate(steps_list):
        step_no = step.get("step") if step.get("step") is not None else i + 1
        expected = step.get("expected") or step.get("expected_result") or ""
        if expected:
            lines.append(f"{step_no}. {expected}")
    
    return "\n\n".join(lines)


def _calculate_row_height(texts: List[str], base: int = 16, max_height: int = 300) -> int:
    """Вычисляет оптимальную высоту строки."""
    max_lines = 1
    for text in texts:
        if text:
            lines = len(str(text).splitlines())
            max_lines = max(max_lines, lines)
    
    # Учитываем двойные переносы между шагами
    calculated = base * max(max_lines, 2) + 10
    return min(calculated, max_height)